from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import re
import openpyxl
import pygame
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui

# Initialize pygame mixer
pygame.mixer.init()

# Load sound file
beep_sound = pygame.mixer.Sound("beep.wav")

# Set volume (0.0 to 1.0)
beep_sound.set_volume(0.1)  # Adjust this value to make the sound quieter

# Function to solve CAPTCHA using Buster
# Function to solve CAPTCHA using Buster
def solve_captcha(driver):
    try:
        # Wait for the CAPTCHA iframe and switch to it
        captcha_iframe = driver.find_element(By.XPATH, "//iframe[contains(@src, 'recaptcha')]")
        driver.switch_to.frame(captcha_iframe)

        # Click inside the CAPTCHA checkbox to open the CAPTCHA widget
        captcha_checkbox = driver.find_element(By.ID, "recaptcha-anchor")
        ActionChains(driver).move_to_element(captcha_checkbox).click().perform()
        time.sleep(2)  # Wait for CAPTCHA challenge to appear

        # Switch back to default content
        driver.switch_to.default_content()

        # Wait for the Buster button to appear (optional, if needed)
        # WebDriverWait(driver, 10).until(
        #     EC.visibility_of_element_located((By.XPATH, "/html/body/div/div/div[3]/div[2]/div[1]/div[1]/div[5]//button"))
        # )

        # Click the Buster button using pyautogui
        pyautogui.click(x=221, y=676)  # Use the specified coordinates
        time.sleep(5)  # Allow time for the CAPTCHA solver to work
        print("Buster CAPTCHA solver activated.")
    except Exception as e:
        print(f"Error activating Buster CAPTCHA solver: {e}")

# Function to get Google search results
def google_search(query):
    # Load Chrome options
    chrome_options = Options()
    chrome_options.add_extension('C:/Users/hamza/Desktop/wordpress/WordPress/Instagram/tools/usernameGoogleSearch/BusterCaptchaSolverforHumans.crx')  # Path to Buster extension .crx file

    driver = webdriver.Chrome(options=chrome_options)  # Ensure the path to chromedriver is correct
    search_results = []
    page = 0

    while True:
        url = f"https://www.google.com/search?q={query}&start={page*10}"
        driver.get(url)
        time.sleep(2)  # Allow time for the page to load

        # Check for CAPTCHA
        if "systems have detected unusual traffic from your com" in driver.page_source:
            print("CAPTCHA detected. Solving with Buster.")
            solve_captcha(driver)  # Call the function to solve CAPTCHA using Buster

        # Extract search result URLs
        links = driver.find_elements(By.CSS_SELECTOR, ".yuRUbf a")
        if not links:  # Exit loop if no more links are found
            break

        for link in links:
            href = link.get_attribute("href")
            if "instagram.com" in href and "/p/" not in href:
                search_results.append(href.split('?')[0])

        page += 1
        time.sleep(2)  # Pause to avoid being blocked

    driver.quit()
    return search_results

# Main function to scrape and save URLs to Excel
def scrape_instagram_urls(keywords_file):
    # Load search keywords from CSV
    df = pd.read_csv(keywords_file, header=None, encoding='latin1')
    keywords = df[0].tolist()

    # Load or create Excel file
    try:
        workbook = openpyxl.load_workbook("instagram_profiles.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Profile URL"])

    # Process each keyword
    for keyword in keywords:
        print(f"Processing keyword: {keyword}")
        search_query = f'{keyword} site:instagram.com intitle:"Instagram photos and videos" -inurl:"/p/" -inurl:"/explore/" -inurl:"help.instagram.com" -inurl:"blog.instagram.com" -inurl:"/tag/"'
        search_results = google_search(search_query)

        # Append new data to the Excel file
        for url in search_results:
            sheet.append([url])
        
        # Save the Excel file after each keyword
        workbook.save("instagram_profiles.xlsx")
        print(f"Data for '{keyword}' saved to instagram_profiles.xlsx")

# Example usage
keywords_file = 'searchKeywords.csv'
scrape_instagram_urls(keywords_file)










######################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import time
import re
import openpyxl
import pygame

# Initialize pygame mixer
pygame.mixer.init()

# Load sound file
beep_sound = pygame.mixer.Sound("beep.wav")

# Set volume (0.0 to 1.0)
beep_sound.set_volume(0.1)  # Adjust this value to make the sound quieter

# Function to get Google search results
def google_search(query):
    driver = webdriver.Chrome()  # Ensure the path to chromedriver is correct
    search_results = []
    page = 0

    while True:
        url = f"https://www.google.com/search?q={query}&start={page*10}"
        driver.get(url)
        time.sleep(2)  # Allow time for the page to load

        # Check for CAPTCHA
        if "systems have detected unusual traffic from your com" in driver.page_source:
            print("CAPTCHA detected. Please resolve it to continue.")
            # Play sound in a loop until CAPTCHA is resolved
            while driver.current_url.startswith("https://www.google.com/sorry/"):
                #beep_sound.play(-1)  # Play sound on loop
                time.sleep(1)
                if not driver.current_url.startswith("https://www.google.com/sorry/"):
                    #beep_sound.stop()  # Stop the sound when CAPTCHA is resolved
                    break
            print("CAPTCHA resolved. Continuing with the search...")

        # Extract search result URLs
        links = driver.find_elements(By.CSS_SELECTOR, ".yuRUbf a")
        if not links:  # Exit loop if no more links are found
            break

        for link in links:
            href = link.get_attribute("href")
            if "instagram.com" in href and "/p/" not in href:
                search_results.append(href.split('?')[0])

        page += 1
        time.sleep(2)  # Pause to avoid being blocked

    driver.quit()
    return search_results

# Main function to scrape and save URLs to Excel
def scrape_instagram_urls(keywords_file):
    # Load search keywords from CSV
    df = pd.read_csv(keywords_file, header=None, encoding='latin1')
    keywords = df[0].tolist()

    # Load or create Excel file
    try:
        workbook = openpyxl.load_workbook("instagram_profiles.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Profile URL"])

    # Process each keyword
    for keyword in keywords:
        print(f"Processing keyword: {keyword}")
        search_query = f'{keyword} site:instagram.com intitle:"Instagram photos and videos" -inurl:"/p/" -inurl:"/explore/" -inurl:"help.instagram.com" -inurl:"blog.instagram.com" -inurl:"/tag/"'
        search_results = google_search(search_query)

        # Append new data to the Excel file
        for url in search_results:
            sheet.append([url])
        
        # Save the Excel file after each keyword
        workbook.save("instagram_profiles.xlsx")
        print(f"Data for '{keyword}' saved to instagram_profiles.xlsx")

# Example usage
keywords_file = 'searchKeywords.csv'
scrape_instagram_urls(keywords_file)










###################################



from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import time
import re
import openpyxl
import pygame

# Initialize pygame mixer
pygame.mixer.init()

# Load sound file
beep_sound = pygame.mixer.Sound("beep.wav")

# Set volume (0.0 to 1.0)
beep_sound.set_volume(0.1)  # Adjust this value to make the sound quieter


# Function to get Google search results
def google_search(driver, query):
    search_results = []
    page = 0

    while True:
        url = f"https://www.google.com/search?q={query}&start={page*10}"
        driver.get(url)
        time.sleep(2)  # Allow time for the page to load

        # Check for CAPTCHA
        if "systems have detected unusual traffic from your com" in driver.page_source:
            print("CAPTCHA detected. Please resolve it to continue.")
            driver.maximize_window()  # Maximize the window when CAPTCHA is detected
            # Play sound in a loop until CAPTCHA is resolved
            while driver.current_url.startswith("https://www.google.com/sorry/"):
                # beep_sound.play(-1)  # Play sound on loop
                time.sleep(1)
                if not driver.current_url.startswith("https://www.google.com/sorry/"):
                    # beep_sound.stop()  # Stop the sound when CAPTCHA is resolved
                    driver.minimize_window()  # Minimize the window after CAPTCHA is resolved
                    break
            print("CAPTCHA resolved. Continuing with the search...")

        # Extract search result URLs
        links = driver.find_elements(By.CSS_SELECTOR, ".yuRUbf a")
        if not links:  # Exit loop if no more links are found
            break

        for link in links:
            href = link.get_attribute("href")
            if "instagram.com" in href and "/p/" not in href:
                search_results.append(href.split('?')[0])

        page += 1
        time.sleep(2)  # Pause to avoid being blocked

    return search_results

# Main function to scrape and save URLs to Excel
def scrape_instagram_urls(keywords_file):
    # Initialize the WebDriver (outside the loop)
    driver = webdriver.Chrome()  # Ensure the path to chromedriver is correct
    driver.minimize_window()  # Minimize the window initially

    # Load search keywords from CSV
    df = pd.read_csv(keywords_file, header=None, encoding='latin1')
    keywords = df[0].tolist()

    # Load or create Excel file
    try:
        workbook = openpyxl.load_workbook("instagram_profiles.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Profile URL"])

    # Process each keyword
    for keyword in keywords:
        print(f"Processing keyword: {keyword}")
        search_query = f'{keyword} site:instagram.com intitle:"Instagram photos and videos" -inurl:"/p/" -inurl:"/explore/" -inurl:"help.instagram.com" -inurl:"blog.instagram.com" -inurl:"/tag/"'
        search_results = google_search(driver, search_query)

        # Append new data to the Excel file
        for url in search_results:
            sheet.append([url])
        
        # Save the Excel file after each keyword
        workbook.save("instagram_profiles.xlsx")
        print(f"Data for '{keyword}' saved to instagram_profiles.xlsx")

    # Quit the browser only after all keywords are processed
    driver.quit()

# Example usage
keywords_file = 'searchKeywords.csv'
scrape_instagram_urls(keywords_file)
